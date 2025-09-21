# -*- coding: utf-8 -*-
"""
تطبيق Flask بسيط لإدارة داخلية لشركة "شركة كثيب للاستثمار".
يحوي نماذج SQLite للمالكين، المستأجرين، المشاريع، الوحدات، و الدفعات.
يتضمن حساب نسبة الشركة واحتساب ضريبة القيمة المضافة.
ملاحظة: هذا تطبيق تعريفي - لتحويله للإنتاج ستحتاج لإضافة مصادقة قوية، صلاحيات، واختبارات.
"""

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_wtf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_compress import Compress
from datetime import datetime, timedelta, timezone
import csv, io, json, os, secrets
from openpyxl import Workbook
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from pymongo import MongoClient
import re

app = Flask(__name__)

# Compression
compress = Compress(app)

# Security Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or secrets.token_hex(32)
app.config['WTF_CSRF_SECRET_KEY'] = os.environ.get('WTF_CSRF_SECRET_KEY') or secrets.token_hex(32)
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

# Database Configuration - Railway provides DATABASE_URL
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Railway provides PostgreSQL
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    # Fallback for local development
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///kthaib_new.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# File Upload Configuration
app.config['UPLOAD_FOLDER'] = 'uploads/contracts'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Security Headers
app.config['SECURITY_HEADERS'] = {
    'X-Content-Type-Options': 'nosniff',
    'X-Frame-Options': 'SAMEORIGIN',
    'X-XSS-Protection': '1; mode=block',
    'Strict-Transport-Security': 'max-age=31536000; includeSubDomains',
    'Content-Security-Policy': "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com; font-src 'self' https://fonts.gstatic.com https://fonts.googleapis.com; img-src 'self' data: https:; connect-src 'self'"
}

# MongoDB Configuration
MONGODB_URL = os.environ.get('MONGODB_URL')
if MONGODB_URL:
    try:
        mongo_client = MongoClient(MONGODB_URL, serverSelectionTimeoutMS=5000)
        mongo_client.admin.command('ping')  # Test connection
        mongo_db = mongo_client['kthaib_db']
        print("MongoDB connected successfully")
    except Exception as e:
        print(f"MongoDB connection failed: {e}")
        mongo_client = None
        mongo_db = None
else:
    print("MONGODB_URL not set, skipping MongoDB")
    mongo_client = None
    mongo_db = None

db = SQLAlchemy(app)

# CSRF Protection
csrf = CSRFProtect(app)

# Rate Limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"]
)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'يرجى تسجيل الدخول للوصول إلى هذه الصفحة.'
login_manager.login_message_category = 'info'
login_manager.session_protection = 'strong'
login_manager.remember_cookie_duration = timedelta(days=30)
login_manager.remember_cookie_secure = True
login_manager.remember_cookie_httponly = True

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.before_request
def check_session_timeout():
    """Check for session timeout and enforce it"""
    if current_user.is_authenticated:
        # Check if session has expired (2 hours of inactivity)
        last_activity = session.get('last_activity')
        if last_activity:
            last_activity_time = datetime.fromisoformat(last_activity).replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) - last_activity_time > timedelta(hours=2):
                logout_user()
                flash('انتهت صلاحية الجلسة بسبب عدم النشاط. يرجى تسجيل الدخول مرة أخرى.', 'info')
                return redirect(url_for('login'))

        # Update last activity time
        session['last_activity'] = datetime.now(timezone.utc).isoformat()

# Security Headers
@app.after_request
def add_security_headers(response):
    """Add security headers to all responses"""
    headers = app.config.get('SECURITY_HEADERS', {})
    for header, value in headers.items():
        response.headers[header] = value
    return response

# HTTPS Enforcement
@app.before_request
def enforce_https():
    """Enforce HTTPS in production"""
    if not request.is_secure and not app.debug:
        url = request.url.replace('http://', 'https://', 1)
        return redirect(url, code=301)

# Input Validation Functions
def validate_email(email):
    """Validate email format"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validate_phone(phone):
    """Validate Saudi phone number format"""
    # Saudi phone numbers: 05xxxxxxxx or +9665xxxxxxxx
    pattern = r'^(\+966|0)?5[0-9]{8}$'
    return re.match(pattern, phone) is not None

def validate_national_id(national_id):
    """Validate Saudi National ID format"""
    # Saudi National ID: 10 digits
    return len(national_id) == 10 and national_id.isdigit()

def sanitize_input(text):
    """Sanitize user input to prevent XSS"""
    if not text:
        return text
    # Remove potentially dangerous characters
    return re.sub(r'[<>]', '', str(text).strip())

def allowed_file(filename):
    """Check if file extension is allowed"""
    ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png', 'gif'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# -------------------------
# نماذج البيانات (Models)
# -------------------------
class Owner(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    national_id = db.Column(db.String(50))
    phone = db.Column(db.String(50))
    email = db.Column(db.String(120))
    address = db.Column(db.String(255))
    sab_number = db.Column(db.String(50))  # رقم الساب
    tenants = db.relationship('Tenant', backref='owner', lazy=True)

# MongoDB Collections
if mongo_db is not None:
    owners_collection = mongo_db['owners']
    tenants_collection = mongo_db['tenants']
    projects_collection = mongo_db['projects']
    units_collection = mongo_db['units']
    payments_collection = mongo_db['payments']
    audit_logs_collection = mongo_db['audit_logs']
else:
    owners_collection = None
    tenants_collection = None
    projects_collection = None
    units_collection = None
    payments_collection = None
    audit_logs_collection = None

class Tenant(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(50))
    contract_start = db.Column(db.Date)
    contract_end = db.Column(db.Date)
    contract_number = db.Column(db.String(100))
    contract_file = db.Column(db.String(255))  # File path for uploaded contract
    sab_number = db.Column(db.String(50))  # رقم الساب
    owner_id = db.Column(db.Integer, db.ForeignKey('owner.id'), nullable=False)  # ربط بالمالك

class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    location = db.Column(db.String(150))
    description = db.Column(db.Text)

class Unit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    unit_number = db.Column(db.String(50))
    type = db.Column(db.String(50))
    area = db.Column(db.Float)
    owner_id = db.Column(db.Integer, db.ForeignKey('owner.id'), nullable=True)
    tenant_id = db.Column(db.Integer, db.ForeignKey('tenant.id'), nullable=True)
    status = db.Column(db.String(50), default='available')

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    unit_id = db.Column(db.Integer, db.ForeignKey('unit.id'), nullable=False)
    payer_type = db.Column(db.String(20))  # owner | tenant
    payer_id = db.Column(db.Integer)
    amount = db.Column(db.Float, nullable=False)
    date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    description = db.Column(db.String(255))
    company_rate = db.Column(db.Float, default=0.05) # نسبة الشركة (قابلة للتعديل عند التسجيل)
    vat_rate = db.Column(db.Float, default=0.15)     # معدل ضريبة القيمة المضافة (قابلة للتعديل)

    # حقول محسوبة تحفظ عند الإنشاء
    company_commission = db.Column(db.Float)
    vat_on_commission = db.Column(db.Float)
    net_to_owner = db.Column(db.Float)

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    action = db.Column(db.String(255))
    user = db.Column(db.String(80))
    timestamp = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # Admin, Accountant, Clerk
    name = db.Column(db.String(120))
    email = db.Column(db.String(120))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    last_login = db.Column(db.DateTime)
    login_attempts = db.Column(db.Integer, default=0)
    locked_until = db.Column(db.DateTime)
    password_reset_token = db.Column(db.String(128))
    password_reset_expires = db.Column(db.DateTime)

    def set_password(self, password):
        # Password strength validation
        if not self.validate_password_strength(password):
            raise ValueError('كلمة المرور ضعيفة. يجب أن تحتوي على 8 أحرف على الأقل، حرف كبير، حرف صغير، ورقم.')
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def validate_password_strength(self, password):
        """Validate password strength requirements"""
        if len(password) < 8:
            return False
        if not any(char.isupper() for char in password):
            return False
        if not any(char.islower() for char in password):
            return False
        if not any(char.isdigit() for char in password):
            return False
        return True

    def is_account_locked(self):
        """Check if account is currently locked"""
        if self.locked_until and datetime.utcnow() < self.locked_until:
            return True
        return False

    def increment_login_attempts(self):
        """Increment login attempts and lock account if necessary"""
        self.login_attempts += 1
        if self.login_attempts >= 5:  # Lock after 5 failed attempts
            self.locked_until = datetime.utcnow() + timedelta(minutes=30)  # Lock for 30 minutes
        db.session.commit()

    def reset_login_attempts(self):
        """Reset login attempts on successful login"""
        self.login_attempts = 0
        self.locked_until = None
        self.last_login = datetime.now(timezone.utc)
        db.session.commit()

    def generate_reset_token(self):
        """Generate password reset token"""
        import secrets
        self.password_reset_token = secrets.token_urlsafe(32)
        self.password_reset_expires = datetime.now(timezone.utc) + timedelta(hours=1)
        db.session.commit()
        return self.password_reset_token

# -------------------------
# وظائف مساعدة
# -------------------------
def calculate_payment_breakdown(amount, company_rate, vat_rate, commission_deducted_from='total'):
    """
    commission_deducted_from:
      - 'total' => العمولة تُحسب من المبلغ الإجمالي (الافتراضي)
      - 'owner' => العمولة تُخصم من المالك (مثال سيناريو مختلف)
    هنا نطبق الصيغة الافتراضية:
    companyCommission = amount * companyRate
    VATonCommission = companyCommission * vatRate
    netToOwner = amount - companyCommission - VATonCommission
    """
    companyCommission = round(amount * company_rate, 2)
    vatOnCommission = round(companyCommission * vat_rate, 2)
    netToOwner = round(amount - companyCommission - vatOnCommission, 2)
    return companyCommission, vatOnCommission, netToOwner

def generate_comprehensive_report(start_date=None, end_date=None, project_id=None, payer_type=None):
    """توليد تقرير نصي شامل محسن للدفعات"""
    query = Payment.query

    # تطبيق الفلاتر
    if start_date:
        query = query.filter(Payment.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Payment.date <= datetime.strptime(end_date, '%Y-%m-%d'))
    if project_id:
        query = query.join(Unit).filter(Unit.project_id == int(project_id))
    if payer_type:
        query = query.filter(Payment.payer_type == payer_type)

    payments = query.order_by(Payment.date.desc()).all()

    # إحصائيات أساسية محسنة
    total_payments = sum(p.amount for p in payments) if payments else 0
    total_commissions = sum(p.company_commission for p in payments) if payments else 0
    total_vat = sum(p.vat_on_commission for p in payments) if payments else 0
    net_to_owners = sum(p.net_to_owner for p in payments) if payments else 0

    # إحصائيات متقدمة
    owner_payments = [p for p in payments if p.payer_type == 'owner']
    tenant_payments = [p for p in payments if p.payer_type == 'tenant']

    # إحصائيات المشاريع
    project_stats = {}
    for payment in payments:
        unit = db.session.get(Unit, payment.unit_id)
        if unit:
            project = db.session.get(Project, unit.project_id)
            project_name = project.name if project else "غير محدد"
            if project_name not in project_stats:
                project_stats[project_name] = {'count': 0, 'total': 0, 'commissions': 0}
            project_stats[project_name]['count'] += 1
            project_stats[project_name]['total'] += payment.amount
            project_stats[project_name]['commissions'] += payment.company_commission

    # إحصائيات شهرية محسنة
    monthly_stats = {}
    quarterly_stats = {}
    for payment in payments:
        month_key = payment.date.strftime('%Y-%m')
        quarter_key = f"{payment.date.year}-Q{(payment.date.month-1)//3 + 1}"

        if month_key not in monthly_stats:
            monthly_stats[month_key] = {'count': 0, 'total': 0, 'commissions': 0, 'owners': 0, 'tenants': 0}
        if quarter_key not in quarterly_stats:
            quarterly_stats[quarter_key] = {'count': 0, 'total': 0, 'commissions': 0}

        monthly_stats[month_key]['count'] += 1
        monthly_stats[month_key]['total'] += payment.amount
        monthly_stats[month_key]['commissions'] += payment.company_commission
        if payment.payer_type == 'owner':
            monthly_stats[month_key]['owners'] += 1
        else:
            monthly_stats[month_key]['tenants'] += 1

        quarterly_stats[quarter_key]['count'] += 1
        quarterly_stats[quarter_key]['total'] += payment.amount
        quarterly_stats[quarter_key]['commissions'] += payment.company_commission

    # إحصائيات الأداء
    avg_payment = total_payments / len(payments) if payments else 0
    avg_commission = total_commissions / len(payments) if payments else 0
    commission_percentage = (total_commissions / total_payments * 100) if total_payments > 0 else 0

    # تحليل الاتجاهات
    trend_analysis = []
    if len(monthly_stats) > 1:
        months = sorted(monthly_stats.keys())
        if len(months) >= 2:
            current_month = months[-1]
            prev_month = months[-2]
            current_total = monthly_stats[current_month]['total']
            prev_total = monthly_stats[prev_month]['total']
            if prev_total > 0:
                growth_rate = ((current_total - prev_total) / prev_total) * 100
                trend_analysis.append(f"معدل النمو الشهري: {growth_rate:+.2f}%")

    # بناء التقرير المحسن
    report = []

    # رأس التقرير
    report.append("╔" + "═" * 78 + "╗")
    report.append("║                    تقرير شامل للدفعات - شركة كثيب للاستثمار                    ║")
    report.append("╚" + "═" * 78 + "╝")
    report.append("")
    report.append(f"📅 تاريخ إنشاء التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report.append(f"📊 عدد السجلات المُحلّلة: {len(payments)}")
    report.append(f"🔄 حالة التقرير: مُحدّث تلقائياً من قاعدة البيانات")
    report.append("")

    # فلاتر المستخدمة
    if start_date or end_date or project_id or payer_type:
        report.append("🔍 الفلاتر المطبقة:")
        report.append("-" * 50)
        if start_date:
            report.append(f"  📅 من تاريخ: {start_date}")
        if end_date:
            report.append(f"  📅 إلى تاريخ: {end_date}")
        if project_id:
            project = db.session.get(Project, int(project_id))
            if project:
                report.append(f"  🏗️  المشروع: {project.name}")
        if payer_type:
            type_name = "👤 مالك" if payer_type == 'owner' else "🏢 مستأجر"
            report.append(f"  {type_name}")
        report.append("")

    # الملخص التنفيذي
    report.append("📈 الملخص التنفيذي:")
    report.append("=" * 50)
    report.append(f"💰 إجمالي المبالغ المدفوعة: {total_payments:,.2f} ريال")
    report.append(f"💼 إجمالي العمولات: {total_commissions:,.2f} ريال ({commission_percentage:.1f}%)")
    report.append(f"🧾 إجمالي ضريبة القيمة المضافة: {total_vat:,.2f} ريال")
    report.append(f"✅ صافي المبالغ للمالكين: {net_to_owners:,.2f} ريال")
    report.append("")
    report.append(f"📊 متوسط المبلغ لكل دفعة: {avg_payment:,.2f} ريال")
    report.append(f"📈 متوسط العمولة لكل دفعة: {avg_commission:,.2f} ريال")
    if trend_analysis:
        report.append(f"📉 {trend_analysis[0]}")
    report.append("")

    # إحصائيات مفصلة حسب النوع
    report.append("👥 تحليل حسب نوع الدافع:")
    report.append("-" * 50)

    owner_count = len(owner_payments)
    tenant_count = len(tenant_payments)
    owner_percentage = (owner_count / len(payments) * 100) if payments else 0
    tenant_percentage = (tenant_count / len(payments) * 100) if payments else 0

    report.append(f"👤 دفعات المالكين: {owner_count} دفعة ({owner_percentage:.1f}%)")
    if owner_payments:
        owner_total = sum(p.amount for p in owner_payments)
        owner_avg = owner_total / owner_count
        report.append(f"   💰 إجمالي: {owner_total:,.2f} ريال")
        report.append(f"   📊 متوسط: {owner_avg:,.2f} ريال")

    report.append(f"🏢 دفعات المستأجرين: {tenant_count} دفعة ({tenant_percentage:.1f}%)")
    if tenant_payments:
        tenant_total = sum(p.amount for p in tenant_payments)
        tenant_avg = tenant_total / tenant_count
        report.append(f"   💰 إجمالي: {tenant_total:,.2f} ريال")
        report.append(f"   📊 متوسط: {tenant_avg:,.2f} ريال")
    report.append("")

    # إحصائيات المشاريع
    if project_stats:
        report.append("🏗️  تحليل حسب المشاريع:")
        report.append("-" * 50)
        for project_name, stats in sorted(project_stats.items(), key=lambda x: x[1]['total'], reverse=True):
            percentage = (stats['total'] / total_payments * 100) if total_payments > 0 else 0
            report.append(f"📁 {project_name}")
            report.append(f"   🔢 عدد الدفعات: {stats['count']}")
            report.append(f"   💰 إجمالي المبالغ: {stats['total']:,.2f} ريال ({percentage:.1f}%)")
            report.append(f"   💼 إجمالي العمولات: {stats['commissions']:,.2f} ريال")
            report.append("")
        report.append("")

    # التحليل الشهري
    if monthly_stats:
        report.append("📅 التحليل الشهري:")
        report.append("-" * 50)
        report.append("شهر        │ دفعات │ مالكين │ مستأجرين │ إجمالي المبالغ │ العمولات")
        report.append("-" * 70)
        for month in sorted(monthly_stats.keys(), reverse=True)[:6]:  # آخر 6 أشهر
            stats = monthly_stats[month]
            report.append(f"{month:<10} │ {stats['count']:<6} │ {stats['owners']:<7} │ {stats['tenants']:<9} │ {stats['total']:>13,.2f} │ {stats['commissions']:>8,.2f}")
        report.append("")

    # التحليل الربع سنوي
    if quarterly_stats:
        report.append("📊 التحليل الربع سنوي:")
        report.append("-" * 50)
        for quarter in sorted(quarterly_stats.keys(), reverse=True):
            stats = quarterly_stats[quarter]
            report.append(f"🗓️  {quarter}")
            report.append(f"   🔢 عدد الدفعات: {stats['count']}")
            report.append(f"   💰 إجمالي المبالغ: {stats['total']:,.2f} ريال")
            report.append(f"   💼 إجمالي العمولات: {stats['commissions']:,.2f} ريال")
            report.append("")

    # مؤشرات الأداء الرئيسية
    report.append("🎯 مؤشرات الأداء الرئيسية (KPIs):")
    report.append("-" * 50)
    report.append(f"📈 معدل العمولة: {commission_percentage:.2f}%")
    report.append(f"💰 متوسط حجم الصفقة: {avg_payment:,.2f} ريال")
    report.append(f"⚡ عدد الصفقات يومياً: {len(payments) / max(1, (datetime.now().date() - payments[-1].date.date()).days + 1) if payments else 0:.1f}")
    report.append(f"🎪 تنوع المشاريع: {len(project_stats)} مشروع")
    report.append("")

    # أكبر الصفقات
    if payments:
        top_payments = sorted(payments, key=lambda x: x.amount, reverse=True)[:5]
        report.append("🏆 أكبر الصفقات:")
        report.append("-" * 50)
        for i, payment in enumerate(top_payments, 1):
            report.append(f"{i}. 💰 {payment.amount:,.2f} ريال")
            report.append(f"   📅 {payment.date.strftime('%Y-%m-%d')}")
            report.append(f"   {'👤 مالك' if payment.payer_type == 'owner' else '🏢 مستأجر'}")
            if payment.description:
                report.append(f"   📝 {payment.description}")
            report.append("")

    # التوصيات
    report.append("💡 توصيات وتحليل:")
    report.append("-" * 50)
    if commission_percentage > 10:
        report.append("⚠️  معدل العمولة مرتفع نسبياً، قد يحتاج إلى مراجعة")
    elif commission_percentage < 3:
        report.append("⚠️  معدل العمولة منخفض، تأكد من الربحية")

    if len(project_stats) > 5:
        report.append("📊 تنوع جيد في المشاريع، مما يقلل من المخاطر")
    elif len(project_stats) <= 2:
        report.append("⚠️  تركيز على عدد محدود من المشاريع، قد يزيد من المخاطر")

    if trend_analysis and trend_analysis[0].startswith("معدل النمو"):
        growth = float(trend_analysis[0].split(": ")[1].rstrip("%"))
        if growth > 10:
            report.append("🚀 نمو ممتاز في الإيرادات")
        elif growth < -5:
            report.append("⚠️  انخفاض في الإيرادات، يحتاج إلى تحليل")
    report.append("")

    # خاتمة التقرير
    report.append("╔" + "═" * 78 + "╗")
    report.append("║                              نهاية التقرير الشامل                              ║")
    report.append("╚" + "═" * 78 + "╝")
    report.append("")
    report.append("📋 تم إنشاء هذا التقرير بواسطة نظام إدارة شركة كثيب للاستثمار")
    report.append("🏢 التقرير يعكس البيانات الحالية في قاعدة البيانات")
    report.append("⚡ تم تحديث التقرير في الوقت الفعلي")
    report.append("")
    report.append("📞 للاستفسارات أو الدعم الفني، يرجى التواصل مع قسم تقنية المعلومات")
    report.append(f"⏰ وقت إنشاء التقرير: {datetime.now().strftime('%H:%M:%S')}")

    return "\n".join(report)

def seed_sample_data():
    """ملئ بيانات تجريبية إذا كانت الجداول فارغة"""
    # SQLite seeding
    if Owner.query.count() == 0:
        owners = [
            Owner(name="محمود العسيري", national_id="1010101010", phone="0500000001", email="ma@kthaib.com", address="الرياض", sab_number="ساب-001"),
            Owner(name="نورة الشمري", national_id="2020202020", phone="0500000002", email="ns@kthaib.com", address="جدة", sab_number="ساب-002"),
            Owner(name="أحمد الخالدي", national_id="3030303030", phone="0500000003", email="ah@kthaib.com", address="الدمام", sab_number="ساب-003"),
            Owner(name="فاطمة الزهراء", national_id="4040404040", phone="0500000004", email="fz@kthaib.com", address="مكة", sab_number="ساب-004"),
            Owner(name="سعد المنصور", national_id="5050505050", phone="0500000005", email="sm@kthaib.com", address="المدينة", sab_number="ساب-005")
        ]
        db.session.add_all(owners)
    if Tenant.query.count() == 0:
        tenants = [
            Tenant(name="شركة الريادة", phone="0590000001", contract_start=datetime(2024,1,1), contract_end=datetime(2026,1,1), sab_number="ساب-ت-001", owner_id=1),
            Tenant(name="مؤسسة النور", phone="0590000002", contract_start=datetime(2024,3,1), contract_end=datetime(2025,3,1), sab_number="ساب-ت-002", owner_id=2),
            Tenant(name="شركة الأمل", phone="0590000003", contract_start=datetime(2024,6,1), contract_end=datetime(2027,6,1), sab_number="ساب-ت-003", owner_id=3)
        ]
        db.session.add_all(tenants)
    if Project.query.count() == 0:
        projects = [
            Project(name="مشروع الريان", location="الرياض", description="مشروع سكني راقٍ"),
            Project(name="مجمع النخيل", location="جدة", description="مجمع تجاري وسكني"),
            Project(name="برج الشروق", location="الدمام", description="برج سكني فاخر")
        ]
        db.session.add_all(projects)
        db.session.flush()
        units = [
            Unit(project_id=1, unit_number="A-101", type="شقة", area=120, owner_id=1, tenant_id=1, status="rented"),
            Unit(project_id=1, unit_number="A-102", type="شقة", area=100, owner_id=2, status="available"),
            Unit(project_id=1, unit_number="A-103", type="شقة", area=150, owner_id=3, status="sold"),
            Unit(project_id=2, unit_number="B-201", type="مكتب", area=80, owner_id=4, tenant_id=2, status="rented"),
            Unit(project_id=2, unit_number="B-202", type="مكتب", area=90, owner_id=5, status="available"),
            Unit(project_id=3, unit_number="C-301", type="شقة", area=200, owner_id=1, tenant_id=3, status="rented")
        ]
        db.session.add_all(units)
    if Payment.query.count() == 0:
        payments = []
        for i in range(1, 7):
            company_rate = 0.05
            vat_rate = 0.15
            amt = 4000 + i * 500  # Vary amounts
            comm, vat, net = calculate_payment_breakdown(amt, company_rate, vat_rate)
            pay = Payment(
                unit_id=i if i <= 6 else 1,
                payer_type='tenant' if i % 2 == 1 else 'owner',
                payer_id=1,
                amount=amt,
                description=f"دفعة تجريبية {i}",
                company_rate=company_rate,
                vat_rate=vat_rate,
                company_commission=comm,
                vat_on_commission=vat,
                net_to_owner=net
            )
            payments.append(pay)
        db.session.add_all(payments)

    # Seed users
    if User.query.count() == 0:
        users = [
            User(username='admin', name='مدير النظام', email='admin@kthaib.com', role='Admin'),
            User(username='accountant', name='المحاسب', email='acc@kthaib.com', role='Accountant'),
            User(username='clerk', name='موظف الاستقبال', email='clerk@kthaib.com', role='Clerk'),
            User(username='manager', name='مدير المشاريع', email='manager@kthaib.com', role='Manager')
        ]
        for user in users:
            # Use stronger default passwords
            default_password = f"{user.username}Admin123!"  # e.g., adminAdmin123!
            user.set_password(default_password)
        db.session.add_all(users)

    db.session.commit()

    # MongoDB seeding
    if mongo_db is not None:
        try:
            if owners_collection.count_documents({}) == 0:
                mongo_owners = [
                    {"name": "محمود العسيري", "national_id": "1010101010", "phone": "0500000001", "email": "ma@kthaib.com", "address": "الرياض", "sab_number": "ساب-001"},
                    {"name": "نورة الشمري", "national_id": "2020202020", "phone": "0500000002", "email": "ns@kthaib.com", "address": "جدة", "sab_number": "ساب-002"},
                    {"name": "أحمد الخالدي", "national_id": "3030303030", "phone": "0500000003", "email": "ah@kthaib.com", "address": "الدمام", "sab_number": "ساب-003"},
                    {"name": "فاطمة الزهراء", "national_id": "4040404040", "phone": "0500000004", "email": "fz@kthaib.com", "address": "مكة", "sab_number": "ساب-004"},
                    {"name": "سعد المنصور", "national_id": "5050505050", "phone": "0500000005", "email": "sm@kthaib.com", "address": "المدينة", "sab_number": "ساب-005"}
                ]
                owners_collection.insert_many(mongo_owners)
                print("MongoDB seeded successfully")
        except Exception as e:
            print(f"MongoDB seeding failed: {e}")
    else:
        print("MongoDB not available, skipping MongoDB seeding")

# -------------------------
# Routes رئيسية
# -------------------------
@app.route('/')
def home():
    return render_template('homepage.html')

@app.route('/health')
def health_check():
    """Health check endpoint for Railway monitoring"""
    return {
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat(),
        'version': '1.0.0'
    }

@app.route('/login', methods=['GET','POST'])
@limiter.limit("5 per minute")
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = sanitize_input(request.form.get('username', ''))
        password = request.form.get('password', '')
        remember = request.form.get('remember') == 'on'

        # Input validation
        if not username or len(username) < 3:
            flash('اسم المستخدم مطلوب ويجب أن يكون 3 أحرف على الأقل.', 'danger')
            return render_template('login.html')

        if not password:
            flash('كلمة المرور مطلوبة.', 'danger')
            return render_template('login.html')

        user = User.query.filter_by(username=username).first()

        if not user:
            flash('اسم المستخدم غير موجود.', 'danger')
            return render_template('login.html')

        if user.is_account_locked():
            remaining_time = int((user.locked_until - datetime.utcnow()).total_seconds() / 60)
            flash(f'الحساب مقفل مؤقتاً. يرجى المحاولة مرة أخرى بعد {remaining_time} دقيقة.', 'warning')
            return render_template('login.html')

        if not user.is_active:
            flash('الحساب غير نشط. يرجى التواصل مع الإدارة.', 'danger')
            return render_template('login.html')

        if user.check_password(password):
            # Successful login
            user.reset_login_attempts()
            login_user(user, remember=remember, duration=timedelta(days=30) if remember else None)

            # Log successful login with IP
            user_ip = request.remote_addr
            db.session.add(AuditLog(action=f"تسجيل دخول ناجح للمستخدم {username} من IP: {user_ip}", user=username))
            db.session.commit()

            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            # Failed login attempt
            user.increment_login_attempts()

            # Log failed login attempt with IP
            user_ip = request.remote_addr
            db.session.add(AuditLog(action=f"محاولة تسجيل دخول فاشلة للمستخدم {username} من IP: {user_ip}", user=username))
            db.session.commit()

            if user.login_attempts >= 5:
                flash('تم قفل الحساب بسبب محاولات تسجيل الدخول المتكررة. يرجى المحاولة مرة أخرى بعد 30 دقيقة.', 'danger')
            else:
                remaining_attempts = 5 - user.login_attempts
                flash(f'كلمة المرور غير صحيحة. لديك {remaining_attempts} محاولات متبقية.', 'danger')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('تم تسجيل الخروج بنجاح.', 'success')
    return redirect(url_for('login'))

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email).first()

        if user:
            # Generate reset token
            reset_token = user.generate_reset_token()
            # In a real application, you would send an email here
            # For demo purposes, we'll show the token
            flash(f'تم إرسال رابط إعادة تعيين كلمة المرور إلى بريدك الإلكتروني. الرمز المؤقت: {reset_token}', 'info')
        else:
            flash('لم يتم العثور على حساب بهذا البريد الإلكتروني.', 'warning')

    return render_template('forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    user = User.query.filter_by(password_reset_token=token).first()

    if not user or (user.password_reset_expires and datetime.utcnow() > user.password_reset_expires):
        flash('رابط إعادة تعيين كلمة المرور غير صالح أو منتهي الصلاحية.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if password != confirm_password:
            flash('كلمة المرور وتأكيدها غير متطابقين.', 'danger')
            return render_template('reset_password.html', token=token)

        try:
            user.set_password(password)
            user.password_reset_token = None
            user.password_reset_expires = None
            db.session.commit()

            # Log password reset
            db.session.add(AuditLog(action=f"إعادة تعيين كلمة المرور للمستخدم {user.username}", user=user.username))
            db.session.commit()

            flash('تم إعادة تعيين كلمة المرور بنجاح. يمكنك الآن تسجيل الدخول.', 'success')
            return redirect(url_for('login'))
        except ValueError as e:
            flash(str(e), 'danger')
            return render_template('reset_password.html', token=token)

    return render_template('reset_password.html', token=token)

@app.route('/dashboard')
@login_required
def dashboard():
    # مؤشرات سريعة (KPIs)
    total_payments = db.session.query(db.func.sum(Payment.amount)).scalar() or 0
    total_commissions = db.session.query(db.func.sum(Payment.company_commission)).scalar() or 0
    total_vat = db.session.query(db.func.sum(Payment.vat_on_commission)).scalar() or 0
    net_paid_to_owners = db.session.query(db.func.sum(Payment.net_to_owner)).scalar() or 0

    # إحصائيات إضافية
    total_owners = Owner.query.count()
    total_tenants = Tenant.query.count()
    total_projects = Project.query.count()
    total_units = Unit.query.count()
    available_units = Unit.query.filter_by(status='available').count()
    rented_units = Unit.query.filter_by(status='rented').count()

    # بيانات لعرض الجداول
    recent_payments = Payment.query.order_by(Payment.date.desc()).limit(10).all()
    recent_audit_logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(10).all()
    projects = Project.query.all()
    owners = Owner.query.all()
    tenants = Tenant.query.all()

    user_role = current_user.role if current_user.is_authenticated else None

    return render_template('dashboard.html',
                           total_payments=total_payments,
                           total_commissions=total_commissions,
                           total_vat=total_vat,
                           net_paid_to_owners=net_paid_to_owners,
                           total_owners=total_owners,
                           total_tenants=total_tenants,
                           total_projects=total_projects,
                           total_units=total_units,
                           available_units=available_units,
                           rented_units=rented_units,
                           recent_payments=recent_payments,
                           recent_audit_logs=recent_audit_logs,
                           projects=projects, owners=owners, tenants=tenants,
                           user_role=user_role,
                           active_page='dashboard')

# إدارة المستخدمين
@app.route('/users', methods=['GET', 'POST'])
@login_required
def users_view():
    if current_user.role != 'Admin':
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        name = request.form.get('name')
        email = request.form.get('email')
        role = request.form.get('role')

        # Validation
        if User.query.filter_by(username=username).first():
            flash('اسم المستخدم موجود بالفعل.', 'danger')
            return redirect(url_for('users_view'))

        if password != confirm_password:
            flash('كلمة المرور وتأكيدها غير متطابقين.', 'danger')
            return redirect(url_for('users_view'))

        try:
            user = User(username=username, name=name, email=email, role=role)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()

            # Log user creation
            db.session.add(AuditLog(action=f"إنشاء مستخدم جديد: {username}", user=current_user.username if current_user.is_authenticated else 'system'))
            db.session.commit()

            flash('تم إضافة المستخدم بنجاح.', 'success')
            return redirect(url_for('users_view'))
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for('users_view'))

    users = User.query.all()
    user_role = current_user.role if current_user.is_authenticated else None
    current_user_name = current_user.name if current_user.is_authenticated else None
    current_username = current_user.username if current_user.is_authenticated else None
    return render_template('users.html', users=users, user_role=user_role, current_user_name=current_user_name, current_username=current_username, active_page='users')

# إدارة الملاك
@app.route('/owners', methods=['GET','POST'])
@login_required
def owners_view():
    if request.method == 'POST':
        # Get and sanitize form data
        name = sanitize_input(request.form.get('name', ''))
        national_id = sanitize_input(request.form.get('national_id', ''))
        phone = sanitize_input(request.form.get('phone', ''))
        email = sanitize_input(request.form.get('email', ''))
        address = sanitize_input(request.form.get('address', ''))
        sab_number = sanitize_input(request.form.get('sab_number', ''))

        # Validation
        if not name or len(name) < 2:
            flash('اسم المالك مطلوب ويجب أن يكون حرفين على الأقل.', 'danger')
            return redirect(url_for('owners_view'))

        if national_id and not validate_national_id(national_id):
            flash('رقم الهوية الوطنية يجب أن يكون 10 أرقام.', 'danger')
            return redirect(url_for('owners_view'))

        if phone and not validate_phone(phone):
            flash('رقم الهاتف يجب أن يكون بتنسيق سعودي صحيح (05xxxxxxxx).', 'danger')
            return redirect(url_for('owners_view'))

        if email and not validate_email(email):
            flash('البريد الإلكتروني غير صحيح.', 'danger')
            return redirect(url_for('owners_view'))

        # Check for duplicate national_id
        if national_id:
            existing_owner = Owner.query.filter_by(national_id=national_id).first()
            if existing_owner:
                flash('رقم الهوية الوطنية موجود بالفعل.', 'danger')
                return redirect(url_for('owners_view'))

        # إنشاء مالك جديد
        o = Owner(
            name=name,
            national_id=national_id,
            phone=phone,
            email=email,
            address=address,
            sab_number=sab_number
        )
        db.session.add(o)
        db.session.commit()
        db.session.add(AuditLog(action=f"إنشاء مالك {o.name}", user=current_user.username if current_user.is_authenticated else 'system'))
        db.session.commit()
        flash('تم إضافة المالك بنجاح.', 'success')
        return redirect(url_for('owners_view'))

    owners = Owner.query.all()
    return render_template('owners.html', owners=owners, active_page='owners')

# إدارة المستأجرين
@app.route('/tenants', methods=['GET','POST'])
def tenants_view():
    if request.method == 'POST':
        owner_id = request.form.get('owner_id')
        if not owner_id:
            flash('يجب اختيار المالك.', 'error')
            return redirect(url_for('tenants_view'))

        owner = db.session.get(Owner, int(owner_id))
        if not owner:
            flash('المالك غير موجود.', 'error')
            return redirect(url_for('tenants_view'))

        if owner.tenants:
            flash('لا يمكن إضافة مستأجر جديد لهذا المالك لأنه يمتلك مستأجر بالفعل. يجب حذف المستأجر الحالي أولاً.', 'error')
            return redirect(url_for('tenants_view'))

        contract_start = request.form.get('contract_start') or None
        contract_end = request.form.get('contract_end') or None
        contract_number = request.form.get('contract_number')
        sab_number = request.form.get('sab_number')

        # Handle file upload with security checks
        contract_file_path = None
        if 'contract_file' in request.files:
            file = request.files['contract_file']
            if file and file.filename:
                # Security checks
                if not allowed_file(file.filename):
                    flash('نوع الملف غير مسموح به. يرجى رفع ملف PDF أو DOC أو صورة فقط.', 'danger')
                    return redirect(url_for('tenants_view'))

                # Check file size (additional check beyond MAX_CONTENT_LENGTH)
                file.seek(0, os.SEEK_END)
                file_size = file.tell()
                file.seek(0)
                if file_size > 10 * 1024 * 1024:  # 10MB limit for contracts
                    flash('حجم الملف كبير جداً. الحد الأقصى 10 ميجابايت.', 'danger')
                    return redirect(url_for('tenants_view'))

                filename = secure_filename(file.filename)
                # Add timestamp and random string to avoid filename conflicts
                import time
                timestamp = str(int(time.time()))
                random_suffix = secrets.token_hex(4)
                filename = f"{timestamp}_{random_suffix}_{filename}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

                # Ensure upload directory exists
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

                file.save(file_path)
                contract_file_path = filename  # Store only the filename, not the full path

        t = Tenant(
            name=request.form.get('name'),
            phone=request.form.get('phone'),
            contract_start=datetime.strptime(contract_start, '%Y-%m-%d') if contract_start else None,
            contract_end=datetime.strptime(contract_end, '%Y-%m-%d') if contract_end else None,
            contract_number=contract_number,
            contract_file=contract_file_path,
            sab_number=sab_number,
            owner_id=int(owner_id)
        )
        db.session.add(t)
        db.session.commit()
        flash('تم إضافة المستأجر.', 'success')
        return redirect(url_for('tenants_view'))
    tenants = Tenant.query.all()
    owners = Owner.query.all()
    return render_template('tenants.html', tenants=tenants, owners=owners, active_page='tenants')

@app.route('/edit_owner/<int:owner_id>', methods=['GET', 'POST'])
def edit_owner(owner_id):
    owner = db.session.get(Owner, owner_id)
    if not owner:
        flash('المالك غير موجود.', 'error')
        return redirect(url_for('owners_view'))

    if request.method == 'POST':
        owner.name = request.form.get('name')
        owner.national_id = request.form.get('national_id')
        owner.phone = request.form.get('phone')
        owner.email = request.form.get('email')
        owner.address = request.form.get('address')
        owner.sab_number = request.form.get('sab_number')
        db.session.commit()
        db.session.add(AuditLog(action=f"تعديل مالك {owner.name}", user='system'))
        db.session.commit()
        flash('تم تحديث المالك بنجاح.', 'success')
        return redirect(url_for('owners_view'))

    return render_template('edit_owner.html', owner=owner)

@app.route('/edit_tenant/<int:tenant_id>', methods=['GET', 'POST'])
def edit_tenant(tenant_id):
    tenant = db.session.get(Tenant, tenant_id)
    if not tenant:
        flash('المستأجر غير موجود.', 'error')
        return redirect(url_for('tenants_view'))

    if request.method == 'POST':
        owner_id = request.form.get('owner_id')
        if not owner_id:
            flash('يجب اختيار المالك.', 'error')
            return redirect(url_for('edit_tenant', tenant_id=tenant_id))

        new_owner_id = int(owner_id)
        if new_owner_id != tenant.owner_id:
            # Changing owner, check if new owner has tenants
            new_owner = db.session.get(Owner, new_owner_id)
            if not new_owner:
                flash('المالك الجديد غير موجود.', 'error')
                return redirect(url_for('edit_tenant', tenant_id=tenant_id))

            if new_owner.tenants:
                flash('لا يمكن نقل المستأجر إلى هذا المالك لأنه يمتلك مستأجر بالفعل. يجب حذف المستأجر الحالي للمالك الجديد أولاً.', 'error')
                return redirect(url_for('edit_tenant', tenant_id=tenant_id))

        tenant.owner_id = new_owner_id
        tenant.name = request.form.get('name')
        tenant.phone = request.form.get('phone')
        tenant.sab_number = request.form.get('sab_number')
        tenant.contract_number = request.form.get('contract_number')

        contract_start = request.form.get('contract_start') or None
        contract_end = request.form.get('contract_end') or None
        tenant.contract_start = datetime.strptime(contract_start, '%Y-%m-%d') if contract_start else None
        tenant.contract_end = datetime.strptime(contract_end, '%Y-%m-%d') if contract_end else None

        # Handle file upload if a new file is provided
        if 'contract_file' in request.files:
            file = request.files['contract_file']
            if file and file.filename:
                filename = secure_filename(file.filename)
                import time
                timestamp = str(int(time.time()))
                filename = f"{timestamp}_{filename}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                tenant.contract_file = filename  # Store only the filename, not the full path

        db.session.commit()
        db.session.add(AuditLog(action=f"تعديل مستأجر {tenant.name}", user='system'))
        db.session.commit()
        flash('تم تحديث المستأجر بنجاح.', 'success')
        return redirect(url_for('tenants_view'))

    owners = Owner.query.all()
    return render_template('edit_tenant.html', tenant=tenant, owners=owners)

@app.route('/delete_owner/<int:owner_id>', methods=['POST'])
def delete_owner(owner_id):
    owner = db.session.get(Owner, owner_id)
    if owner:
        # Check if owner has tenants
        if owner.tenants:
            return {'success': False, 'message': 'لا يمكن حذف المالك لأنه مرتبط بمستأجرين'}, 400

        db.session.delete(owner)
        db.session.add(AuditLog(action=f"حذف مالك {owner.name}", user='system'))
        db.session.commit()
        return {'success': True}, 200
    return {'success': False}, 404

@app.route('/delete_tenant/<int:tenant_id>', methods=['POST'])
def delete_tenant(tenant_id):
    tenant = db.session.get(Tenant, tenant_id)
    if tenant:
        # Delete contract file if exists
        if tenant.contract_file:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], tenant.contract_file)
            if os.path.exists(file_path):
                os.remove(file_path)

        db.session.delete(tenant)
        db.session.add(AuditLog(action=f"حذف مستأجر {tenant.name}", user='system'))
        db.session.commit()
        return {'success': True}, 200
    return {'success': False}, 404

@app.route('/uploads/contracts/<filename>')
def download_contract(filename):
    # Extract filename from path if it contains directory separators
    clean_filename = filename.split('\\')[-1].split('/')[-1]
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], clean_filename), as_attachment=True)

@app.route('/serve_contract/<filename>')
def serve_contract(filename):
    """Serve contract file inline for viewing in browser"""
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        return "File not found", 404

    file_ext = filename.lower().split('.')[-1]
    mime_types = {
        'pdf': 'application/pdf',
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'gif': 'image/gif'
    }

    mime_type = mime_types.get(file_ext, 'application/octet-stream')
    return send_file(file_path, mimetype=mime_type, as_attachment=False)

@app.route('/view_contract/<filename>')
def view_contract(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        flash('الملف غير موجود.', 'error')
        return redirect(url_for('tenants_view'))

    # Get file information
    file_ext = filename.lower().split('.')[-1]
    file_stats = os.stat(file_path)
    upload_date = datetime.fromtimestamp(file_stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S')

    # Determine if file can be viewed inline
    viewable_types = ['pdf', 'jpg', 'jpeg', 'png', 'gif']

    if file_ext in viewable_types:
        # Render the view template for supported file types
        return render_template('view_contract.html',
                             filename=filename,
                             file_type=file_ext,
                             upload_date=upload_date)
    else:
        # For unsupported files, redirect to download
        flash('نوع الملف غير مدعوم للعرض المباشر. سيتم تحميل الملف.', 'info')
        return redirect(url_for('download_contract', filename=filename))

# إدارة المشاريع والوحدات
@app.route('/projects', methods=['GET','POST'])
def projects_view():
    if request.method == 'POST':
        name = request.form.get('name')
        location = request.form.get('location')
        description = request.form.get('description')

        # Validation
        if not name or len(name.strip()) == 0:
            flash('يجب إدخال اسم المشروع.', 'error')
            return redirect(url_for('projects_view'))

        # Check if project name already exists
        existing_project = Project.query.filter_by(name=name.strip()).first()
        if existing_project:
            flash('اسم المشروع موجود بالفعل.', 'error')
            return redirect(url_for('projects_view'))

        try:
            p = Project(name=name.strip(), location=location, description=description)
            db.session.add(p)
            db.session.commit()
            flash('تم إضافة المشروع بنجاح.', 'success')
        except Exception as e:
            flash('حدث خطأ أثناء إضافة المشروع.', 'error')

        return redirect(url_for('projects_view'))

    # Handle search and filters
    search = request.args.get('search', '')
    status_filter = request.args.get('status_filter', '')
    project_filter = request.args.get('project_filter', '')

    # Filter projects
    projects_query = Project.query
    if search:
        projects_query = projects_query.filter(
            db.or_(
                Project.name.contains(search),
                Project.location.contains(search),
                Project.description.contains(search)
            )
        )
    projects = projects_query.all()

    # Filter and paginate units
    units_query = Unit.query
    if status_filter:
        units_query = units_query.filter(Unit.status == status_filter)
    if project_filter:
        units_query = units_query.filter(Unit.project_id == int(project_filter))

    # Pagination for units
    page = int(request.args.get('page', 1))
    per_page = 10
    units_paginated = units_query.paginate(page=page, per_page=per_page, error_out=False)
    units = units_paginated.items
    total_pages = units_paginated.pages

    owners = Owner.query.all()
    tenants = Tenant.query.all()
    return render_template('projects.html', projects=projects, units=units, owners=owners, tenants=tenants,
                           page=page, total_pages=total_pages, active_page='projects')

@app.route('/add_unit', methods=['POST'])
def add_unit():
    try:
        project_id = request.form.get('project_id')
        unit_number = request.form.get('unit_number')
        area = request.form.get('area')

        # Validation
        if not project_id:
            flash('يجب اختيار المشروع.', 'error')
            return redirect(url_for('projects_view'))

        if not unit_number or len(unit_number.strip()) == 0:
            flash('يجب إدخال رقم الوحدة.', 'error')
            return redirect(url_for('projects_view'))

        # Check if unit number already exists in the project
        existing_unit = Unit.query.filter_by(project_id=int(project_id), unit_number=unit_number.strip()).first()
        if existing_unit:
            flash('رقم الوحدة موجود بالفعل في هذا المشروع.', 'error')
            return redirect(url_for('projects_view'))

        if area and float(area) <= 0:
            flash('يجب أن تكون المساحة أكبر من صفر.', 'error')
            return redirect(url_for('projects_view'))

        unit = Unit(
            project_id=int(project_id),
            unit_number=unit_number.strip(),
            type=request.form.get('type'),
            area=float(area) if area else 0,
            owner_id=int(request.form.get('owner_id')) if request.form.get('owner_id') else None,
            status=request.form.get('status') or 'available'
        )
        db.session.add(unit)
        db.session.commit()
        flash('تم إضافة الوحدة بنجاح.', 'success')
    except ValueError as e:
        flash('خطأ في البيانات المدخلة.', 'error')
    except Exception as e:
        flash('حدث خطأ غير متوقع.', 'error')

    return redirect(url_for('projects_view'))

@app.route('/edit_project/<int:project_id>', methods=['GET', 'POST'])
def edit_project(project_id):
    project = db.session.get(Project, project_id)
    if not project:
        flash('المشروع غير موجود.', 'error')
        return redirect(url_for('projects_view'))

    if request.method == 'POST':
        project.name = request.form.get('name')
        project.location = request.form.get('location')
        project.description = request.form.get('description')
        db.session.commit()
        db.session.add(AuditLog(action=f"تعديل مشروع {project.name}", user='system'))
        db.session.commit()
        flash('تم تحديث المشروع بنجاح.', 'success')
        return redirect(url_for('projects_view'))

    return render_template('edit_project.html', project=project)

@app.route('/delete_project/<int:project_id>', methods=['POST'])
def delete_project(project_id):
    project = db.session.get(Project, project_id)
    if not project:
        flash('المشروع غير موجود.', 'error')
        return redirect(url_for('projects_view'))

    # Check if project has units
    units_count = Unit.query.filter_by(project_id=project_id).count()
    if units_count > 0:
        flash(f'لا يمكن حذف المشروع لأنه يحتوي على {units_count} وحدة.', 'error')
        return redirect(url_for('projects_view'))

    db.session.delete(project)
    db.session.add(AuditLog(action=f"حذف مشروع {project.name}", user='system'))
    db.session.commit()
    flash('تم حذف المشروع بنجاح.', 'success')
    return redirect(url_for('projects_view'))

@app.route('/edit_unit/<int:unit_id>', methods=['GET', 'POST'])
def edit_unit(unit_id):
    unit = db.session.get(Unit, unit_id)
    if not unit:
        flash('الوحدة غير موجودة.', 'error')
        return redirect(url_for('projects_view'))

    projects = Project.query.all()
    owners = Owner.query.all()
    tenants = Tenant.query.all()

    if request.method == 'POST':
        unit.project_id = int(request.form.get('project_id'))
        unit.unit_number = request.form.get('unit_number')
        unit.type = request.form.get('type')
        unit.area = float(request.form.get('area') or 0)
        unit.owner_id = int(request.form.get('owner_id')) if request.form.get('owner_id') else None
        unit.tenant_id = int(request.form.get('tenant_id')) if request.form.get('tenant_id') else None
        unit.status = request.form.get('status')
        db.session.commit()
        db.session.add(AuditLog(action=f"تعديل وحدة {unit.unit_number}", user='system'))
        db.session.commit()
        flash('تم تحديث الوحدة بنجاح.', 'success')
        return redirect(url_for('projects_view'))

    return render_template('edit_unit.html', unit=unit, projects=projects, owners=owners, tenants=tenants)

@app.route('/delete_unit/<int:unit_id>', methods=['POST'])
def delete_unit(unit_id):
    unit = db.session.get(Unit, unit_id)
    if not unit:
        flash('الوحدة غير موجودة.', 'error')
        return redirect(url_for('projects_view'))

    # Check if unit has payments
    payments_count = Payment.query.filter_by(unit_id=unit_id).count()
    if payments_count > 0:
        flash(f'لا يمكن حذف الوحدة لأنها تحتوي على {payments_count} دفعة.', 'error')
        return redirect(url_for('projects_view'))

    db.session.delete(unit)
    db.session.add(AuditLog(action=f"حذف وحدة {unit.unit_number}", user='system'))
    db.session.commit()
    flash('تم حذف الوحدة بنجاح.', 'success')
    return redirect(url_for('projects_view'))

# تسجيل دفعات
@app.route('/payments', methods=['GET','POST'])
@login_required
@limiter.limit("100 per hour")
def payments_view():
    units = Unit.query.all()
    owners_query = Owner.query.all()
    tenants_query = Tenant.query.all()
    owners = [{'id': o.id, 'name': o.name, 'national_id': o.national_id} for o in owners_query]
    tenants = [{'id': t.id, 'name': t.name} for t in tenants_query]
    projects = Project.query.all()

    # فلترة الدفعات
    query = Payment.query
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    project_id = request.args.get('project_id')
    search = request.args.get('search', '')

    if start_date:
        query = query.filter(Payment.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Payment.date <= datetime.strptime(end_date, '%Y-%m-%d'))
    if project_id:
        query = query.join(Unit).filter(Unit.project_id == int(project_id))
    if search:
        # البحث في الوصف أو المبلغ
        query = query.filter(
            db.or_(
                Payment.description.contains(search),
                Payment.amount.cast(db.String).contains(search)
            )
        )

    # Pagination
    page = int(request.args.get('page', 1))
    per_page = 10
    payments_paginated = query.order_by(Payment.date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    payments = payments_paginated.items
    total_pages = payments_paginated.pages

    # Create dictionaries for efficient lookup
    owners_dict = {owner.id: owner for owner in owners_query}
    tenants_dict = {tenant.id: tenant for tenant in tenants_query}

    # Add payer names to payments for efficient template rendering
    for payment in payments:
        if payment.payer_type == 'owner' and payment.payer_id in owners_dict:
            payment.payer_name = owners_dict[payment.payer_id].name
        elif payment.payer_type == 'tenant' and payment.payer_id in tenants_dict:
            payment.payer_name = tenants_dict[payment.payer_id].name
        else:
            payment.payer_name = 'غير محدد'
    if request.method == 'POST':
        unit_id = int(request.form.get('unit_id'))
        payer_type = request.form.get('payer_type')
        payer_id = int(request.form.get('payer_id')) if request.form.get('payer_id') else None
        amount = float(request.form.get('amount') or 0)
        company_rate = float(request.form.get('company_rate') or 0.05)
        vat_rate = float(request.form.get('vat_rate') or 0.15)
        payment_date = request.form.get('payment_date')
        if payment_date:
            payment_date = datetime.strptime(payment_date, '%Y-%m-%d')
        else:
            payment_date = datetime.now(timezone.utc)

        # التحقق من صحة payer_id
        if payer_type == 'owner':
            payer = db.session.get(Owner, payer_id)
            if not payer:
                flash('معرف المالك غير صحيح.', 'error')
                return redirect(url_for('payments_view'))
        elif payer_type == 'tenant':
            payer = db.session.get(Tenant, payer_id)
            if not payer:
                flash('معرف المستأجر غير صحيح.', 'error')
                return redirect(url_for('payments_view'))

        # حساب القيم
        comm, vat, net = calculate_payment_breakdown(amount, company_rate, vat_rate)
        p = Payment(unit_id=unit_id, payer_type=payer_type, payer_id=payer_id, amount=amount,
                    description=request.form.get('description'),
                    company_rate=company_rate, vat_rate=vat_rate,
                    company_commission=comm, vat_on_commission=vat, net_to_owner=net,
                    date=payment_date)
        db.session.add(p)
        db.session.add(AuditLog(action=f"تسجيل دفعة للوحدة {unit_id} مبلغ {amount}", user='system'))
        db.session.commit()
        flash('تم تسجيل الدفعة بنجاح.', 'success')
        return redirect(url_for('payments_view'))
    return render_template('payments.html', units=units, owners=owners, tenants=tenants, payments=payments, projects=projects,
                           page=page, total_pages=total_pages, search=search, active_page='payments')

@app.route('/delete_payment/<int:payment_id>', methods=['DELETE'])
def delete_payment(payment_id):
    payment = db.session.get(Payment, payment_id)
    if payment:
        db.session.delete(payment)
        db.session.add(AuditLog(action=f"حذف دفعة {payment_id}", user='system'))
        db.session.commit()
        return {'success': True}, 200
    return {'success': False}, 404

@app.route('/edit_payment/<int:payment_id>', methods=['GET', 'POST'])
def edit_payment(payment_id):
    payment = db.session.get(Payment, payment_id)
    if not payment:
        flash('الدفعة غير موجودة.', 'error')
        return redirect(url_for('payments_view'))

    units = Unit.query.all()
    owners = [{'id': o.id, 'name': o.name, 'national_id': o.national_id} for o in Owner.query.all()]
    tenants = [{'id': t.id, 'name': t.name} for t in Tenant.query.all()]

    if request.method == 'POST':
        unit_id = int(request.form.get('unit_id'))
        payer_type = request.form.get('payer_type')
        payer_id = int(request.form.get('payer_id')) if request.form.get('payer_id') else None
        amount = float(request.form.get('amount') or 0)
        company_rate = float(request.form.get('company_rate') or 0.05)
        vat_rate = float(request.form.get('vat_rate') or 0.15)
        payment_date = request.form.get('payment_date')
        if payment_date:
            payment_date = datetime.strptime(payment_date, '%Y-%m-%d')
        else:
            payment_date = payment.date

        # Validation
        if payer_type == 'owner':
            payer = db.session.get(Owner, payer_id)
            if not payer:
                flash('معرف المالك غير صحيح.', 'error')
                return redirect(url_for('edit_payment', payment_id=payment_id))
        elif payer_type == 'tenant':
            payer = db.session.get(Tenant, payer_id)
            if not payer:
                flash('معرف المستأجر غير صحيح.', 'error')
                return redirect(url_for('edit_payment', payment_id=payment_id))

        # Update payment
        payment.unit_id = unit_id
        payment.payer_type = payer_type
        payment.payer_id = payer_id
        payment.amount = amount
        payment.company_rate = company_rate
        payment.vat_rate = vat_rate
        payment.date = payment_date
        payment.description = request.form.get('description')

        # Recalculate
        comm, vat, net = calculate_payment_breakdown(amount, company_rate, vat_rate)
        payment.company_commission = comm
        payment.vat_on_commission = vat
        payment.net_to_owner = net

        db.session.commit()
        db.session.add(AuditLog(action=f"تعديل دفعة {payment_id}", user='system'))
        db.session.commit()
        flash('تم تحديث الدفعة بنجاح.', 'success')
        return redirect(url_for('payments_view'))

    return render_template('edit_payment.html', payment=payment, units=units, owners=owners, tenants=tenants)

# تقرير وتصدير
@app.route('/reports')
def reports_view():
    # فلترة بسيطة ممكن توسيعها بواسطة باراميترات GET (project, owner, date range...)
    query = Payment.query

    # فلاتر
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    project_id = request.args.get('project_id')
    payer_type = request.args.get('payer_type')

    if start_date:
        query = query.filter(Payment.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Payment.date <= datetime.strptime(end_date, '%Y-%m-%d'))
    if project_id:
        # فلترة حسب المشروع عبر الوحدات
        query = query.join(Unit).filter(Unit.project_id == int(project_id))
    if payer_type:
        query = query.filter(Payment.payer_type == payer_type)

    payments = query.order_by(Payment.date.desc()).all()

    # ملخصات
    total_payments = sum(p.amount for p in payments) if payments else 0
    total_commissions = sum(p.company_commission for p in payments) if payments else 0
    total_vat = sum(p.vat_on_commission for p in payments) if payments else 0
    net_to_owners = sum(p.net_to_owner for p in payments) if payments else 0

    # قائمة المشاريع للفلترة
    projects = Project.query.all()

    # توليد التقرير النصي الشامل
    text_report = generate_comprehensive_report(start_date, end_date, project_id, payer_type)

    return render_template('reports.html', payments=payments, total_payments=total_payments,
                           total_commissions=total_commissions, total_vat=total_vat,
                           net_to_owners=net_to_owners, projects=projects, text_report=text_report,
                           active_page='reports')

@app.route('/export/payments/csv')
def export_payments_csv():
    query = Payment.query

    # نفس الفلاتر من reports_view
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    project_id = request.args.get('project_id')
    payer_type = request.args.get('payer_type')

    if start_date:
        query = query.filter(Payment.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Payment.date <= datetime.strptime(end_date, '%Y-%m-%d'))
    if project_id:
        query = query.join(Unit).filter(Unit.project_id == int(project_id))
    if payer_type:
        query = query.filter(Payment.payer_type == payer_type)

    payments = query.order_by(Payment.date.desc()).all()

    si = io.StringIO()
    cw = csv.writer(si)
    cw.writerow(["id","unit_id","payer_type","amount","date","company_commission","vat_on_commission","net_to_owner","description"])
    for p in payments:
        cw.writerow([p.id,p.unit_id,p.payer_type,p.amount,p.date.strftime("%Y-%m-%d %H:%M:%S"),p.company_commission,p.vat_on_commission,p.net_to_owner,p.description])
    output = io.BytesIO()
    output.write(si.getvalue().encode('utf-8'))
    output.seek(0)
    return send_file(output, mimetype='text/csv', download_name='payments.csv', as_attachment=True)

@app.route('/export/payments/excel')
def export_payments_excel():
    query = Payment.query

    # نفس الفلاتر من reports_view
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    project_id = request.args.get('project_id')
    payer_type = request.args.get('payer_type')

    if start_date:
        query = query.filter(Payment.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Payment.date <= datetime.strptime(end_date, '%Y-%m-%d'))
    if project_id:
        query = query.join(Unit).filter(Unit.project_id == int(project_id))
    if payer_type:
        query = query.filter(Payment.payer_type == payer_type)

    payments = query.order_by(Payment.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "الدفعات"

    # Headers
    headers = ["ID", "الوحدة", "نوع الدافع", "المبلغ", "التاريخ", "عمولة الشركة", "ضريبة القيمة المضافة", "صافي للمالك", "الوصف"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Data
    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1, value=payment.id)
        ws.cell(row=row_num, column=2, value=payment.unit_id)
        ws.cell(row=row_num, column=3, value=payment.payer_type)
        ws.cell(row=row_num, column=4, value=payment.amount)
        ws.cell(row=row_num, column=5, value=payment.date.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_num, column=6, value=payment.company_commission)
        ws.cell(row=row_num, column=7, value=payment.vat_on_commission)
        ws.cell(row=row_num, column=8, value=payment.net_to_owner)
        ws.cell(row=row_num, column=9, value=payment.description)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name='payments.xlsx', as_attachment=True)

@app.route('/api/unit/<int:unit_id>')
def get_unit_details(unit_id):
    unit = db.session.get(Unit, unit_id)
    if not unit:
        return {'error': 'Unit not found'}, 404

    project = db.session.get(Project, unit.project_id) if unit.project_id else None

    unit_data = {
        'id': unit.id,
        'unit_number': unit.unit_number,
        'project_name': project.name if project else '',
        'status': unit.status,
        'is_rented': unit.tenant_id is not None
    }

    if unit.tenant_id:
        tenant = db.session.get(Tenant, unit.tenant_id)
        if tenant:
            unit_data['tenant'] = {
                'id': tenant.id,
                'name': tenant.name,
                'contract_start': tenant.contract_start.strftime('%Y-%m-%d') if tenant.contract_start else None,
                'contract_end': tenant.contract_end.strftime('%Y-%m-%d') if tenant.contract_end else None
            }
            # Assuming rent amount is not stored, but perhaps we can use a default or from contract
            # For now, leave rent_amount as None
            unit_data['rent_amount'] = None
            unit_data['rent_date'] = tenant.contract_start.strftime('%Y-%m-%d') if tenant.contract_start else None

    return unit_data

@app.route('/export/payments/text')
def export_payments_text():
    # الحصول على الفلاتر من الطلب
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    project_id = request.args.get('project_id')
    payer_type = request.args.get('payer_type')

    # توليد التقرير النصي الشامل
    report_text = generate_comprehensive_report(start_date, end_date, project_id, payer_type)

    # تحسين التنسيق للتصدير - إضافة رأس وتذييل باللغة العربية
    export_header = "=" * 100 + "\n"
    export_header += "ملف التقرير الشامل - شركة كثيب للاستثمار\n"
    export_header += "تم التصدير في: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + "\n"
    export_header += "نوع الملف: تقرير نصي شامل\n"
    export_header += "=" * 100 + "\n\n"

    export_footer = "\n" + "=" * 100 + "\n"
    export_footer += "نهاية الملف المُصدّر\n"
    export_footer += "تم إنشاء هذا التقرير بواسطة نظام إدارة شركة كثيب للاستثمار\n"
    export_footer += "جميع الحقوق محفوظة © شركة كثيب للاستثمار\n"
    export_footer += "=" * 100 + "\n"

    # دمج الرأس والتقرير والتذييل
    full_report = export_header + report_text + export_footer

    # إرسال الملف مع ترميز UTF-8 لضمان عرض النص العربي بشكل صحيح
    output = io.BytesIO()
    output.write(full_report.encode('utf-8-sig'))  # استخدام utf-8-sig لدعم العربية في Windows
    output.seek(0)

    # اسم الملف باللغة العربية مع التاريخ
    filename = f"تقرير_شامل_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    return send_file(
        output,
        mimetype='text/plain; charset=utf-8',
        download_name=filename,
        as_attachment=True
    )

# -------------------------
# تشغيل التطبيق وتهيئة DB
# -------------------------
if __name__ == '__main__':
    with app.app_context():
        # Always recreate database for development (remove this in production)
        db.create_all()
        seed_sample_data()

        # Create upload folder if it doesn't exist
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'])

    # For Railway deployment
    port = int(os.environ.get('PORT', 8000))
    app.run(host='127.0.0.1', port=port, debug=False)